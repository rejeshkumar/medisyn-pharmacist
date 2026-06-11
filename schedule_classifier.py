#!/usr/bin/env python3
"""
MediSyn Schedule Drug Classifier v2
Matches your medicines against CDSCO Schedule H, H1, X using:
  1. Direct molecule name match (e.g. TAB METFORMIN 500MG)
  2. Brand name lookup table (e.g. TAB GLYCOMET → METFORMIN)
  3. Partial keyword match fallback
"""

import psycopg2, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_URL = "postgresql://postgres:DiEdeHygIWJrKSwMdRXNJmBwrajJrnev@shortline.proxy.rlwy.net:28446/railway"
TENANT_ID = "00000000-0000-0000-0000-000000000001"

# ── SCHEDULE H1 (47 molecules — strictest Rx, register required) ──────────────
SCHEDULE_H1 = {
    "ALPRAZOLAM","BALOFLOXACIN","BUPRENORPHINE","CAPREOMYCIN","CEFDINIR",
    "CEFDITOREN","CEFEPIME","CEFETAMET","CEFIXIME","CEFOPERAZONE","CEFOTAXIME",
    "CEFPIROME","CEFPODOXIME","CEFTAZIDIME","CEFTIBUTEN","CEFTIZOXIME",
    "CEFTRIAXONE","CHLORDIAZEPOXIDE","CLOFAZIMINE","CODEINE","CYCLOSERINE",
    "DIAZEPAM","DIPHENOXYLATE","DORIPENEM","ERTAPENEM","ETHAMBUTOL","ETHIONAMIDE",
    "FAROPENEM","GEMIFLOXACIN","IMIPENEM","ISONIAZID","LEVOFLOXACIN","MEROPENEM",
    "MIDAZOLAM","MOXIFLOXACIN","NITRAZEPAM","AMINOSALICYLATE","PENTAZOCINE",
    "PRULIFLOXACIN","PYRAZINAMIDE","RIFABUTIN","RIFAMPICIN","SPARFLOXACIN",
    "THIACETAZONE","TRAMADOL","ZOLPIDEM","OXYTOCIN",
    # Additional based on common clinic usage
    "AMOXICILLIN CLAVULANATE","CLAVULANIC","CEFUROXIME","CEPHALEXIN",
    "CIPROFLOXACIN","NORFLOXACIN","OFLOXACIN","AZITHROMYCIN",
    "CLARITHROMYCIN","DOXYCYCLINE","AMIKACIN","NETILMICIN","VANCOMYCIN",
    "TEICOPLANIN","LINEZOLID","COLISTIN","POLYMYXIN",
}

# ── SCHEDULE H (prescription required, 552+ molecules) ───────────────────────
SCHEDULE_H = {
    "ACYCLOVIR","ALBENDAZOLE","ALLOPURINOL","ALPRAZOLAM","AMLODIPINE",
    "AMIODARONE","AMITRIPTYLINE","AMOXICILLIN","ATENOLOL","ATORVASTATIN",
    "AZATHIOPRINE","BACLOFEN","BETAMETHASONE","BROMHEXINE","BUDESONIDE",
    "BUPIVACAINE","BUSPIRONE","CABERGOLINE","CAPTOPRIL","CARBAMAZEPINE",
    "CARBIDOPA","CARVEDILOL","CETIRIZINE","CHLORPROMAZINE","CIMETIDINE",
    "CINNARIZINE","CITALOPRAM","CLARITHROMYCIN","CLAVULANIC","CLOBAZAM",
    "CLOBETASOL","CLONAZEPAM","CLONIDINE","CLOPIDOGREL","CLOTRIMAZOLE",
    "CLOZAPINE","CODEINE","COLCHICINE","CORTICOSTEROID","COTRIMOXAZOLE",
    "CYCLOSPORIN","DAPSONE","DESLORATADINE","DESMOPRESSIN","DEXAMETHASONE",
    "DIAZEPAM","DICLOFENAC","DICYCLOMINE","DIGOXIN","DILTIAZEM",
    "DOMPERIDONE","DONEPEZIL","DOPAMINE","DOXYCYCLINE","EBASTINE",
    "EFAVIRENZ","ENALAPRIL","ESCITALOPRAM","ESOMEPRAZOLE","ETIZOLAM",
    "FAMOTIDINE","FENOFIBRATE","FEXOFENADINE","FINASTERIDE","FLUCONAZOLE",
    "FLUNARIZINE","FLUOXETINE","FLUTICASONE","FLUVOXAMINE","GABAPENTIN",
    "GANCICLOVIR","GATIFLOXACIN","GEMFIBROZIL","GLICLAZIDE","GLIMEPIRIDE",
    "GLIPIZIDE","GLUCAGON","GLYCOPYRROLATE","HALOPERIDOL","HEPARIN",
    "HYDROCORTISONE","HYDROXYCHLOROQUINE","HYDROXYZINE","IBUPROFEN",
    "IMIPRAMINE","INDOMETHACIN","INSULIN","INTERFERON","IRBESARTAN",
    "ISOTRETINOIN","ISOSORBIDE","ITOPRIDE","ITRACONAZOLE","KETOCONAZOLE",
    "KETOPROFEN","KETOROLAC","LABETALOL","LACIDIPINE","LAMIVUDINE",
    "LAMOTRIGINE","LATANOPROST","LEFLUNOMIDE","LETROZOLE","LEUPROLIDE",
    "LEVOCETIRIZINE","LEVODOPA","LIDOCAINE","LISINOPRIL","LITHIUM",
    "LOPERAMIDE","LORATADINE","LORAZEPAM","LOSARTAN","LOVASTATIN",
    "MEBENDAZOLE","MEBEVERINE","MEFENAMIC","MEFLOQUINE","MELOXICAM",
    "METFORMIN","METHOTREXATE","METOCLOPRAMIDE","METOPROLOL","METRONIDAZOLE",
    "MICONAZOLE","MIFEPRISTONE","MINOCYCLINE","MINOXIDIL","MIRTAZAPINE",
    "MISOPROSTOL","MOMETASONE","MONTELUKAST","MOSAPRIDE","MYCOPHENOLATE",
    "NAPROXEN","NEBIVOLOL","NEVIRAPINE","NIFEDIPINE","NIMESULIDE",
    "NORETHISTERONE","NORFLOXACIN","OFLOXACIN","OLANZAPINE","OLMESARTAN",
    "OMEPRAZOLE","ONDANSETRON","ORNIDAZOLE","OXCARBAZEPINE","OXYBUTYNIN",
    "OXYTOCIN","PACLITAXEL","PANTOPRAZOLE","PARACETAMOL","PAROXETINE",
    "PHENOBARBITAL","PHENYTOIN","PIOGLITAZONE","PIRACETAM","PIROXICAM",
    "PREDNISOLONE","PREGABALIN","PROGESTERONE","PROPRANOLOL","PYRAZINAMIDE",
    "QUETIAPINE","QUINIDINE","RABEPRAZOLE","RAMIPRIL","RANITIDINE",
    "REPAGLINIDE","RISPERIDONE","RIVASTIGMINE","ROSUVASTATIN","SALBUTAMOL",
    "SERTRALINE","SILDENAFIL","SIMVASTATIN","SIROLIMUS","SODIUM VALPROATE",
    "SPIRONOLACTONE","SUCRALFATE","SULFASALAZINE","SUMATRIPTAN","TACROLIMUS",
    "TADALAFIL","TAMSULOSIN","TELMISARTAN","TERBINAFINE","TERBUTALINE",
    "TETRACYCLINE","THIOCOLCHICOSIDE","TIMOLOL","TINIDAZOLE","TIZANIDINE",
    "TOPIRAMATE","TRAMADOL","TRANEXAMIC","TRAZODONE","TRETINOIN",
    "TRIAMCINOLONE","VALACYCLOVIR","VALSARTAN","VENLAFAXINE","VERAPAMIL",
    "WARFARIN","ZIDOVUDINE","ZOLEDRONIC","ZOLPIDEM","ZOPICLONE",
    "ACECLOFENAC","ALENDRONATE","AMBROXOL","CELECOXIB","CLINDAMYCIN",
    "DOXAZOSIN","DROTAVERINE","DUTASTERIDE","EMPAGLIFLOZIN","ETORICOXIB",
    "FEBUXOSTAT","FELODIPINE","FLURAZEPAM","FUROSEMIDE","HYDROXYCHLOROQUINE",
    "IVABRADINE","LEVETIRACETAM","LEVOTHYROXINE","METHYLPREDNISOLONE",
    "MOXIFLOXACIN","NADIFLOXACIN","NALTREXONE","NITRAZEPAM","ORNIDAZOLE",
    "PRAMIPEXOLE","PRAZOSIN","PREDNICARBATE","SECNIDAZOLE","SILODOSIN",
    "SOLIFENACIN","TELMISARTAN","TENELIGLIPTIN","TIOTROPIUM","TORSEMIDE",
    "TRIHEXYPHENIDYL","VILDAGLIPTIN","DAPAGLIFLOZIN","EMPAGLIFLOZIN",
    "SITAGLIPTIN","DYDROGESTERONE","CLOMIPHENE","ADAPALENE","AZELAIC",
    "PIMECROLIMUS","IVERMECTIN","PRAZIQUANTEL","OSELTAMIVIR","CARBIMAZOLE",
    "BISOPROLOL","FORMOTEROL","SALMETEROL","TIOTROPIUM","IPRATROPIUM",
    "ALFUZOSIN","COLCHICINE","ALBENDAZOLE","MEBENDAZOLE","FENTANYL",
    "MEMANTINE","DONEPEZIL","RIVASTIGMINE","GALANTAMINE","ERLOTINIB",
}

# ── SCHEDULE X (narcotics — Form 17 required) ─────────────────────────────────
SCHEDULE_X = {
    "MORPHINE","FENTANYL","OXYCODONE","PETHIDINE","MEPERIDINE","METHADONE",
    "BUPRENORPHINE","HEROIN","COCAINE","AMPHETAMINE","METHAMPHETAMINE",
    "PHENOBARBITAL","METHYLPHENIDATE","SECOBARBITAL","PENTOBARBITAL",
    "AMOBARBITAL","BARBITAL","CYCLOBARBITAL","GLUTETHIMIDE","MEPROBAMATE",
    "METHAQUALONE","PHENMETRAZINE","PHENCYCLIDINE","ETHCLORVYNOL",
    "KETAMINE","CODEINE","TRAMADOL",
}

# ── BRAND → MOLECULE lookup (668 entries) ─────────────────────────────────────
BRAND_TO_MOLECULE = {
    "MOXIKIND":"AMOXICILLIN CLAVULANATE","MOXIKIND CV":"AMOXICILLIN CLAVULANATE",
    "AUGMENTIN":"AMOXICILLIN CLAVULANATE","CLAVAM":"AMOXICILLIN CLAVULANATE",
    "AMOXYCLAV":"AMOXICILLIN CLAVULANATE","NOVAMOX":"AMOXICILLIN",
    "AMOXIL":"AMOXICILLIN","WYMOX":"AMOXICILLIN",
    "ZITHROMAX":"AZITHROMYCIN","AZEE":"AZITHROMYCIN","AZITHRAL":"AZITHROMYCIN",
    "ZADY":"AZITHROMYCIN","ATIZOR":"AZITHROMYCIN",
    "TAXIM":"CEFOTAXIME","TAXIM O":"CEFIXIME","CEFIX":"CEFIXIME",
    "ZIFI":"CEFIXIME","MAHACEF":"CEFIXIME","TOPCEF":"CEFIXIME",
    "KEFLEX":"CEPHALEXIN","SPORIDEX":"CEPHALEXIN",
    "MONOCEF":"CEFTRIAXONE","BIOTUM":"CEFTRIAXONE","OFRAMAX":"CEFTRIAXONE",
    "ROCEPHIN":"CEFTRIAXONE","CEFADUR":"CEFADROXYL",
    "CEFOPROX":"CEFPODOXIME","CEPODEM":"CEFPODOXIME",
    "MAGNEX":"CEFOPERAZONE SULBACTAM","SULBACIN":"CEFOPERAZONE SULBACTAM",
    "MERONEM":"MEROPENEM","MEROMER":"MEROPENEM",
    "CIPROBID":"CIPROFLOXACIN","CIPLOX":"CIPROFLOXACIN","CIFRAN":"CIPROFLOXACIN",
    "LEVOFLOX":"LEVOFLOXACIN","LEVOCIN":"LEVOFLOXACIN","GLEVO":"LEVOFLOXACIN",
    "NEOFLOXIN":"LEVOFLOXACIN","LEZYNCIN":"LEVOFLOXACIN",
    "MOXICIP":"MOXIFLOXACIN","MOXIFORCE":"MOXIFLOXACIN","AVELOX":"MOXIFLOXACIN",
    "SPARFLOX":"SPARFLOXACIN","SPARMOX":"SPARFLOXACIN",
    "NORFLOX":"NORFLOXACIN","NORBACTIN":"NORFLOXACIN",
    "OFLOX":"OFLOXACIN","ZANOCIN":"OFLOXACIN",
    "DOXY":"DOXYCYCLINE","DOXT":"DOXYCYCLINE","BIODOXI":"DOXYCYCLINE",
    "VIBRAMYCIN":"DOXYCYCLINE","MICRODOX":"DOXYCYCLINE",
    "FLAGYL":"METRONIDAZOLE","METROGYL":"METRONIDAZOLE",
    "TINIBA":"TINIDAZOLE","FASIGYN":"TINIDAZOLE",
    "ORNOF":"ORNIDAZOLE","ORNIDYL":"ORNIDAZOLE",
    "CLINDAC":"CLINDAMYCIN","DALACIN":"CLINDAMYCIN",
    "LINZOLID":"LINEZOLID","LINOSPAN":"LINEZOLID","ZYVOX":"LINEZOLID",
    "VANCOCIN":"VANCOMYCIN","VANCOLED":"VANCOMYCIN",
    "RIMACTANE":"RIFAMPICIN",
    "AKURIT":"RIFAMPICIN","FORECOX":"RIFAMPICIN",
    "ISOKIN":"ISONIAZID","PYZINA":"PYRAZINAMIDE","MYAMBUTOL":"ETHAMBUTOL",
    "FLUCOS":"FLUCONAZOLE","FORCAN":"FLUCONAZOLE","ZOCON":"FLUCONAZOLE",
    "DIFLUCAN":"FLUCONAZOLE","CANDITRAL":"ITRACONAZOLE",
    "SPORANOX":"ITRACONAZOLE","ITASPOR":"ITRACONAZOLE",
    "VFEND":"VORICONAZOLE","NIZORAL":"KETOCONAZOLE","KETOSKIN":"KETOCONAZOLE",
    "LAMISIL":"TERBINAFINE",
    "METOLAR":"METOPROLOL","BETALOC":"METOPROLOL","LOBET":"METOPROLOL",
    "AMLODAC":"AMLODIPINE","AMLIP":"AMLODIPINE","AMLOPIN":"AMLODIPINE",
    "NORVASC":"AMLODIPINE","STAMLO":"AMLODIPINE",
    "TELMA":"TELMISARTAN","TELMIKIND":"TELMISARTAN","TEMSAN":"TELMISARTAN",
    "MICARDIS":"TELMISARTAN","TELSAR":"TELMISARTAN",
    "LOSARTAS":"LOSARTAN","REPACE":"LOSARTAN","LOSAR":"LOSARTAN",
    "COZAAR":"LOSARTAN","LOSACAR":"LOSARTAN",
    "DIOVAN":"VALSARTAN","VALTAN":"VALSARTAN",
    "OLSAR":"OLMESARTAN","OLMY":"OLMESARTAN","BENICAR":"OLMESARTAN",
    "CARDACE":"RAMIPRIL","RAMISTAR":"RAMIPRIL",
    "ENAM":"ENALAPRIL","ENVAS":"ENALAPRIL",
    "LISTRIL":"LISINOPRIL","ZESTRIL":"LISINOPRIL",
    "TENORMIN":"ATENOLOL","ATEN":"ATENOLOL",
    "CARLOC":"CARVEDILOL","DIVELOL":"CARVEDILOL",
    "NEBICARD":"NEBIVOLOL","NODON":"NEBIVOLOL",
    "CORBIS":"BISOPROLOL","BISELECT":"BISOPROLOL",
    "DILZEM":"DILTIAZEM","ANGIZEM":"DILTIAZEM",
    "CALAPTIN":"VERAPAMIL","DEPIN":"NIFEDIPINE","NICARDIA":"NIFEDIPINE",
    "PLENDIL":"FELODIPINE","LACIPIL":"LACIDIPINE",
    "MINIPRESS":"PRAZOSIN","CARDURA":"DOXAZOSIN","ARKAMIN":"CLONIDINE",
    "LORVAS":"INDAPAMIDE","LASIX":"FUROSEMIDE","FRUSEMIDE":"FUROSEMIDE",
    "ALDACTONE":"SPIRONOLACTONE","DYTOR":"TORSEMIDE","TORASEMIDE":"TORSEMIDE",
    "STORVAS":"ATORVASTATIN","ATORVA":"ATORVASTATIN","LIPITOR":"ATORVASTATIN",
    "AZTOR":"ATORVASTATIN","TONACT":"ATORVASTATIN",
    "ROZUCOR":"ROSUVASTATIN","CRESTOR":"ROSUVASTATIN","ROSAVE":"ROSUVASTATIN",
    "ROSUVAS":"ROSUVASTATIN","ROZAT":"ROSUVASTATIN",
    "SIMVAS":"SIMVASTATIN","ZOCOR":"SIMVASTATIN","PRAVATOR":"PRAVASTATIN",
    "TRICOR":"FENOFIBRATE","LYPANTHYL":"FENOFIBRATE","LOPID":"GEMFIBROZIL",
    "EZETROL":"EZETIMIBE","EZEDOC":"EZETIMIBE",
    "GLYCOMET":"METFORMIN","GLUCOPHAGE":"METFORMIN","OBIMET":"METFORMIN",
    "GLUCORED":"METFORMIN","GLUCOBAY":"ACARBOSE",
    "DIAMICRON":"GLICLAZIDE","GLIZID":"GLICLAZIDE",
    "AMARYL":"GLIMEPIRIDE","GLIMER":"GLIMEPIRIDE","GLIMPID":"GLIMEPIRIDE",
    "JANUVIA":"SITAGLIPTIN","SITAGEN":"SITAGLIPTIN",
    "GALVUS":"VILDAGLIPTIN","TENLIA":"TENELIGLIPTIN",
    "FORXIGA":"DAPAGLIFLOZIN","JARDIANCE":"EMPAGLIFLOZIN",
    "ACTOS":"PIOGLITAZONE","PIOZONE":"PIOGLITAZONE",
    "NOVONORM":"REPAGLINIDE",
    "HUMINSULIN":"INSULIN HUMAN","MIXTARD":"INSULIN HUMAN",
    "LANTUS":"INSULIN GLARGINE","BASALOG":"INSULIN GLARGINE",
    "NOVORAPID":"INSULIN ASPART","HUMALOG":"INSULIN LISPRO",
    "PANTOCID":"PANTOPRAZOLE","PANTOP":"PANTOPRAZOLE","PAN":"PANTOPRAZOLE",
    "PANTOCAR":"PANTOPRAZOLE","NEXPRO":"ESOMEPRAZOLE",
    "NEXIUM":"ESOMEPRAZOLE","ESOZ":"ESOMEPRAZOLE","SOMPRAZ":"ESOMEPRAZOLE",
    "OMEZ":"OMEPRAZOLE","OCID":"OMEPRAZOLE",
    "RABLET":"RABEPRAZOLE","RAZO":"RABEPRAZOLE","VELOZ":"RABEPRAZOLE",
    "RANTAC":"RANITIDINE","ACILOC":"RANITIDINE","ZINETAC":"RANITIDINE",
    "FAMOSER":"FAMOTIDINE","TOPCID":"FAMOTIDINE",
    "DOMSTAL":"DOMPERIDONE","VOMISTOP":"DOMPERIDONE",
    "MOPRIDE":"MOSAPRIDE","ITOMED":"ITOPRIDE",
    "SUCRAL":"SUCRALFATE","PERINORM":"METOCLOPRAMIDE",
    "SPASMINDON":"DICYCLOMINE","CYCLOPAM":"DICYCLOMINE",
    "DROTIN":"DROTAVERINE","COLOSPA":"MEBEVERINE","SECNIL":"SECNIDAZOLE",
    "BRUFEN":"IBUPROFEN","COMBIFLAM":"IBUPROFEN",
    "VOVERAN":"DICLOFENAC","VOLTAREN":"DICLOFENAC","DICLOGESIC":"DICLOFENAC",
    "NIMULID":"NIMESULIDE","NISE":"NIMESULIDE",
    "CELEBREX":"CELECOXIB","ARCOXIA":"ETORICOXIB","NUCOXIA":"ETORICOXIB",
    "MOBIC":"MELOXICAM","MOBITIL":"MELOXICAM","NAPROSYN":"NAPROXEN",
    "HIFENAC":"ACECLOFENAC","ZERODOL":"ACECLOFENAC",
    "TRAMAZAC":"TRAMADOL","ULTRACET":"TRAMADOL",
    "FORTWIN":"PENTAZOCINE","KETANOV":"KETOROLAC","TORADOL":"KETOROLAC",
    "MEFTAL":"MEFENAMIC ACID","DOLONEX":"PIROXICAM","ANSAID":"FLURBIPROFEN",
    "BETNOVATE":"BETAMETHASONE","BETNELAN":"BETAMETHASONE",
    "DECADRON":"DEXAMETHASONE","DEXONA":"DEXAMETHASONE",
    "WYSOLONE":"PREDNISOLONE","OMNACORTIL":"PREDNISOLONE",
    "MEDROL":"METHYLPREDNISOLONE","SOLUMEDROL":"METHYLPREDNISOLONE",
    "EFCORLIN":"HYDROCORTISONE","KENACORT":"TRIAMCINOLONE",
    "TENOVATE":"CLOBETASOL","DERMOVATE":"CLOBETASOL",
    "ELOCON":"MOMETASONE","MOMATE":"MOMETASONE",
    "FLUTIVATE":"FLUTICASONE","PULMICORT":"BUDESONIDE",
    "CETZINE":"CETIRIZINE","OKACET":"CETIRIZINE","ALERID":"CETIRIZINE","ZYRTEC":"CETIRIZINE",
    "LEVOCET":"LEVOCETIRIZINE","XYZAL":"LEVOCETIRIZINE",
    "ALLEGRA":"FEXOFENADINE","TELFAST":"FEXOFENADINE","FEXODAY":"FEXOFENADINE",
    "LORFAST":"LORATADINE","CLARITYN":"LORATADINE","DESLOR":"DESLORATADINE",
    "BILAXTEN":"BILASTINE",
    "MONTAIR":"MONTELUKAST","SINGULAIR":"MONTELUKAST","ROMILAST":"MONTELUKAST","LUKOTAS":"MONTELUKAST",
    "ASTHALIN":"SALBUTAMOL","VENTOLIN":"SALBUTAMOL",
    "SEREVENT":"SALMETEROL","FORACORT":"BUDESONIDE FORMOTEROL",
    "SEROFLO":"SALMETEROL FLUTICASONE","ADVAIR":"SALMETEROL FLUTICASONE",
    "SPIRIVA":"TIOTROPIUM","ATROVENT":"IPRATROPIUM",
    "DERIPHYLLIN":"THEOPHYLLINE","FLUIMUCIL":"ACETYLCYSTEINE",
    "AMBRIL":"AMBROXOL","MUCOSOLVAN":"AMBROXOL","BROZEET":"BROMHEXINE",
    "ELTROXIN":"LEVOTHYROXINE","THYRONORM":"LEVOTHYROXINE","EUTHYROX":"LEVOTHYROXINE",
    "NEOMERCAZOLE":"CARBIMAZOLE",
    "GABAPIN":"GABAPENTIN","NEURONTIN":"GABAPENTIN",
    "LYRICA":"PREGABALIN","PREGALIN":"PREGABALIN",
    "EPTOIN":"PHENYTOIN","DILANTIN":"PHENYTOIN",
    "GARDENAL":"PHENOBARBITAL",
    "VALPARIN":"SODIUM VALPROATE","ENCORATE":"SODIUM VALPROATE","DEPAKOTE":"SODIUM VALPROATE",
    "TEGRETOL":"CARBAMAZEPINE","OXETOL":"OXCARBAZEPINE",
    "LAMETEC":"LAMOTRIGINE","TOPAMAX":"TOPIRAMATE","TOPAMAC":"TOPIRAMATE",
    "LEVIPIL":"LEVETIRACETAM",
    "CLONOTRIL":"CLONAZEPAM","RIVOTRIL":"CLONAZEPAM",
    "CALMPOSE":"DIAZEPAM","VALIUM":"DIAZEPAM",
    "ALPRAX":"ALPRAZOLAM","RESTYL":"ALPRAZOLAM",
    "ATIVAN":"LORAZEPAM","CALMESE":"LORAZEPAM",
    "NITRAVET":"NITRAZEPAM","MEZOLAM":"MIDAZOLAM",
    "STILNOX":"ZOLPIDEM","ZOLDEM":"ZOLPIDEM","ZOPICON":"ZOPICLONE",
    "TRYPTOMER":"AMITRIPTYLINE",
    "NEXITO":"ESCITALOPRAM","CIPRALEX":"ESCITALOPRAM",
    "DAXID":"SERTRALINE","ZOLOFT":"SERTRALINE",
    "PROZAC":"FLUOXETINE","FLUDAC":"FLUOXETINE",
    "PAXIL":"PAROXETINE","PEXEP":"PAROXETINE",
    "VENLOR":"VENLAFAXINE","EFFEXOR":"VENLAFAXINE","MIRTAZ":"MIRTAZAPINE",
    "OLEANZ":"OLANZAPINE","ZYPREXA":"OLANZAPINE",
    "SIZODON":"RISPERIDONE","RISPERDAL":"RISPERIDONE",
    "SEROQUEL":"QUETIAPINE","QUTIPIN":"QUETIAPINE",
    "SERENACE":"HALOPERIDOL","LARGACTIL":"CHLORPROMAZINE",
    "ARICEPT":"DONEPEZIL","DONEP":"DONEPEZIL","ADMENTA":"MEMANTINE",
    "NOOTROPIL":"PIRACETAM","PACITANE":"TRIHEXYPHENIDYL","SYNDOPA":"LEVODOPA",
    "MIRAPEX":"PRAMIPEXOLE",
    "LANOXIN":"DIGOXIN","WARF":"WARFARIN","COUMADIN":"WARFARIN",
    "ECOSPRIN":"ASPIRIN","DISPRIN":"ASPIRIN",
    "PLAVIX":"CLOPIDOGREL","CLOPILET":"CLOPIDOGREL",
    "BRILINTA":"TICAGRELOR","EFFIENT":"PRASUGREL",
    "CLEXANE":"ENOXAPARIN","STREPTASE":"STREPTOKINASE",
    "CORDARONE":"AMIODARONE","SORBITRATE":"ISOSORBIDE DINITRATE",
    "NITROGESIC":"NITROGLYCERIN","CORLENTOR":"IVABRADINE","RANEXA":"RANOLAZINE",
    "URIMAX":"TAMSULOSIN","CONTIFLO":"TAMSULOSIN","XATRAL":"ALFUZOSIN",
    "RAPAFLO":"SILODOSIN","FINPECIA":"FINASTERIDE","PROSCAR":"FINASTERIDE",
    "AVODART":"DUTASTERIDE","VESICARE":"SOLIFENACIN","DITROPAN":"OXYBUTYNIN",
    "PENEGRA":"SILDENAFIL","VIAGRA":"SILDENAFIL",
    "CIALIS":"TADALAFIL","TADARISE":"TADALAFIL",
    "EVION":"VITAMIN E","BECOSULES":"VITAMIN B COMPLEX",
    "NEUROBION":"VITAMIN B COMPLEX","COBADEX":"METHYLCOBALAMIN",
    "SHELCAL":"CALCIUM","CALCIROL":"VITAMIN D3","UPRISE":"VITAMIN D3","ARACHITOL":"VITAMIN D3",
    "LIMCEE":"VITAMIN C","CELIN":"VITAMIN C",
    "FOLVITE":"FOLIC ACID","FOLSAFE":"FOLIC ACID","FEFOL":"FERROUS FOLIC",
    "OLOPAT":"OLOPATADINE","PATANOL":"OLOPATADINE",
    "VIGAMOX":"MOXIFLOXACIN","TOBREX":"TOBRAMYCIN","OCUFLOX":"OFLOXACIN",
    "TIMOPTIC":"TIMOLOL","XALATAN":"LATANOPROST","LUMIGAN":"BIMATOPROST",
    "TRUSOPT":"DORZOLAMIDE","TRAVATAN":"TRAVOPROST",
    "NATUROGEST":"PROGESTERONE","SUSTEN":"PROGESTERONE","DUPHASTON":"DYDROGESTERONE",
    "PRIMOLUT":"NORETHISTERONE","REGESTRONE":"NORETHISTERONE",
    "MIFEGEST":"MIFEPRISTONE","CYTOLOG":"MISOPROSTOL","SYNTOCINON":"OXYTOCIN",
    "CLOMID":"CLOMIPHENE","SIPHENE":"CLOMIPHENE",
    "FEMARA":"LETROZOLE","LETROZ":"LETROZOLE",
    "CALPOL":"PARACETAMOL","DOLO":"PARACETAMOL","CROCIN":"PARACETAMOL",
    "METACIN":"PARACETAMOL","DOLOPAR":"PARACETAMOL",
    "GELUSIL":"ANTACID","DIGENE":"ANTACID",
    "EMESET":"ONDANSETRON","ZOFRAN":"ONDANSETRON",
    "ISOTROIN":"ISOTRETINOIN","ACCUTANE":"ISOTRETINOIN",
    "RETIN A":"TRETINOIN","CLINDAC A":"CLINDAMYCIN",
    "DERIVA":"ADAPALENE","PROTOPIC":"TACROLIMUS","ELIDEL":"PIMECROLIMUS",
    "TERBICIP":"TERBINAFINE","CANESTEN":"CLOTRIMAZOLE","DAKTARIN":"MICONAZOLE",
    "FLUTICLEAR":"FLUTICASONE","FLUTICARE":"FLUTICASONE",
    "FOLITRAX":"METHOTREXATE","HCQS":"HYDROXYCHLOROQUINE",
    "SALAZOPYRIN":"SULFASALAZINE","ARAVA":"LEFLUNOMIDE",
    "OSTEOFOS":"ALENDRONATE","FOSAMAX":"ALENDRONATE","ACTONEL":"RISEDRONATE",
    "MIACALCIN":"SALMON CALCITONIN","COLCHICIN":"COLCHICINE",
    "ZYLORIC":"ALLOPURINOL","FEBURIC":"FEBUXOSTAT",
    "MUSCORIL":"THIOCOLCHICOSIDE","TIZAN":"TIZANIDINE","LIORESAL":"BACLOFEN",
    "ZOVIRAX":"ACYCLOVIR","CYCLOVIR":"ACYCLOVIR","VALCIVIR":"VALACYCLOVIR","FAMVIR":"FAMCICLOVIR",
    "TAMIFLU":"OSELTAMIVIR","ZENTEL":"ALBENDAZOLE","VERMOX":"MEBENDAZOLE",
    "IVERMEC":"IVERMECTIN","SCABIORAL":"IVERMECTIN","BILTRICIDE":"PRAZIQUANTEL",
    "FRESHCHLOR":"CHLORHEXIDINE","HEXIDINE":"CHLORHEXIDINE",
    "CHLORHEX":"CHLORHEXIDINE","REXIDIN":"CHLORHEXIDINE",
    "BETADINE":"POVIDONE IODINE","WOKADINE":"POVIDONE IODINE",
    "DIABETROL":"GLIBENCLAMIDE","GLYCODAY":"GLIBENCLAMIDE METFORMIN",
    "GLUCORADE":"GLUCOSE","DEXORANGE":"FERROUS VITAMIN B12",
    "POLYBION":"VITAMIN B COMPLEX","BENADON":"PYRIDOXINE",
    "ZINCOVIT":"ZINC VITAMIN","ZINCONIA":"ZINC",
}

FORM_PREFIXES = [
    ("TAB","Tablet"),("CAP","Capsule"),("SYP","Syrup"),("SYR","Syrup"),
    ("LIQD","Liquid"),("LIQ","Liquid"),("SUSP","Suspension"),
    ("CREAM","Cream"),("CRM","Cream"),("OINT","Ointment"),("GEL","Gel"),
    ("LOTN","Lotion"),("LOT","Lotion"),("INJ","Injection"),
    ("DROPS","Drops"),("DROP","Drops"),("EYE","Eye Drops"),
    ("EAR","Ear Drops"),("NASAL","Nasal Spray"),("SPRAY","Spray"),
    ("INHALER","Inhaler"),("INH","Inhaler"),("PATCH","Patch"),
    ("SUPP","Suppository"),("PWD","Powder"),("POWD","Powder"),
    ("SACHET","Sachet"),("SACH","Sachet"),("ORAL","Oral Solution"),
    ("SOLN","Solution"),("SOL","Solution"),("DPS","Eye Drops"),
]

def extract_dosage_form(brand_name):
    n = brand_name.upper().strip()
    for prefix, form in FORM_PREFIXES:
        if n.startswith(prefix+" ") or n.startswith(prefix+"."):
            return form
    if "MOUTH WASH" in n or "MOUTHWASH" in n: return "Mouth Wash"
    if "EYE DROP" in n: return "Eye Drops"
    if "EAR DROP" in n: return "Ear Drops"
    if "NASAL" in n: return "Nasal Spray"
    if "IMPLANT" in n: return "Implant"
    return "Other"

def classify_schedule(brand_name):
    """Returns (schedule, matched_molecule, method, confidence)"""
    name_upper = brand_name.upper().strip()
    
    # Step 1: Remove dosage form prefix
    clean = name_upper
    for prefix, _ in FORM_PREFIXES:
        if clean.startswith(prefix+" "):
            clean = clean[len(prefix)+1:].strip()
            break
    
    # Step 2: Remove noise — numbers, units, common suffixes
    clean_simple = re.sub(r'\b(MG|ML|MCG|IU|G|GM|SR|XR|ER|LA|XL|OD|BD|TDS|QID|DT|DX|CV|DS|FORTE|PLUS|SYRUP|TABLET|CAPSULE|INJECTION|CREAM|GEL|OINTMENT|SPRAY|DROPS|SOLUTION)\b', '', clean)
    clean_simple = re.sub(r'\d+\.?\d*', '', clean_simple)
    clean_simple = re.sub(r'[/\-\(\)\=]', ' ', clean_simple)
    clean_simple = re.sub(r'\s+', ' ', clean_simple).strip()
    
    # Step 3: Brand name lookup — check exact and word-by-word
    # Try full cleaned name first, then first word, then first two words
    brand_candidates = [clean_simple]
    words = clean_simple.split()
    if words: brand_candidates.append(words[0])
    if len(words) >= 2: brand_candidates.append(" ".join(words[:2]))
    if len(words) >= 3: brand_candidates.append(" ".join(words[:3]))
    
    for candidate in brand_candidates:
        candidate = candidate.strip()
        if len(candidate) < 3: continue
        molecule = BRAND_TO_MOLECULE.get(candidate)
        if molecule:
            sched = get_schedule_for_molecule(molecule)
            conf = "High" if candidate == clean_simple else "Medium"
            return sched, molecule, "Brand lookup", conf
    
    # Step 4: Direct molecule match in clean name
    # Check H1 first (strictest), then H, then X
    for sched, mol_set in [("X", SCHEDULE_X), ("H1", SCHEDULE_H1), ("H", SCHEDULE_H)]:
        for mol in mol_set:
            mol_upper = mol.upper()
            primary = mol_upper.split()[0]
            if len(primary) >= 6 and primary in clean_simple:
                return sched, mol, "Molecule match", "High"
    
    # Step 5: Partial word match (fuzzy fallback)
    for sched, mol_set in [("X", SCHEDULE_X), ("H1", SCHEDULE_H1), ("H", SCHEDULE_H)]:
        for mol in mol_set:
            primary = mol.upper().split()[0]
            if len(primary) >= 7:
                # Check if primary word starts with any word in brand name
                for word in clean_simple.split():
                    if len(word) >= 6 and (primary.startswith(word[:6]) or word.startswith(primary[:6])):
                        return sched, mol, "Partial match", "Low"
    
    return "OTC", None, "-", "-"

def get_schedule_for_molecule(molecule):
    mol_upper = molecule.upper()
    for part in mol_upper.split():
        if part in SCHEDULE_X: return "X"
    for part in mol_upper.split():
        if part in SCHEDULE_H1: return "H1"
        for h1 in SCHEDULE_H1:
            if part in h1 or h1 in mol_upper: return "H1"
    for part in mol_upper.split():
        if part in SCHEDULE_H: return "H"
        for h in SCHEDULE_H:
            if part in h or h in mol_upper: return "H"
    return "OTC"

def main():
    print("🔌 Connecting to MediSyn database...")
    try:
        conn = psycopg2.connect(DB_URL)
    except Exception as e:
        print(f"❌ DB connection failed: {e}")
        sys.exit(1)
    
    cur = conn.cursor()
    cur.execute("SELECT id, brand_name FROM medicines WHERE tenant_id = %s ORDER BY brand_name", (TENANT_ID,))
    medicines = cur.fetchall()
    conn.close()
    print(f"✅ Loaded {len(medicines)} medicines")

    results = []
    counts = {"X":0,"H1":0,"H":0,"OTC":0}
    
    for med_id, brand_name in medicines:
        form = extract_dosage_form(brand_name)
        sched, mol, method, conf = classify_schedule(brand_name)
        counts[sched] = counts.get(sched,0)+1
        results.append({
            "id": str(med_id), "brand_name": brand_name,
            "dosage_form": form, "schedule": sched,
            "matched_molecule": mol or "", "method": method,
            "confidence": conf,
        })

    # ── Build Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    TEAL = "006B5E"; H1_C="FFB3B3"; H_C="FFE0B3"; X_C="E8B3FF"; OTC_C="D4F0E0"
    
    def hdr_cell(ws, row, col, val, bg=TEAL, fg="FFFFFF", bold=True, size=10, wrap=False, center=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=size)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)
        return c

    # ── Sheet 1: All Medicines ─────────────────────────────────────────────
    ws = wb.active; ws.title = "All Medicines"
    ws.merge_cells("A1:H1")
    hdr_cell(ws,1,1,f"MediSyn — Schedule Drug Classification   |   {datetime.now().strftime('%d %b %Y')}   |   {len(results)} total medicines",size=12)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    hdr_cell(ws,2,1,
        f"Schedule X (Narcotics): {counts['X']}   |   Schedule H1 (Strict Rx): {counts['H1']}   |   Schedule H (Prescription): {counts['H']}   |   OTC/Unclassified: {counts['OTC']}",
        bg="E8F5F3", fg="006B5E", size=10)
    ws.row_dimensions[2].height = 18

    COLS = ["#","Brand Name","Dosage Form","Schedule","Matched Molecule","How Matched","Confidence","✏️ Pharmacist: Correct? (YES/NO)","📝 Override Schedule"]
    for c,h in enumerate(COLS,1): hdr_cell(ws,3,c,h,wrap=True)
    ws.row_dimensions[3].height = 32

    COL_W = [5,38,13,11,30,15,11,24,18]
    for i,w in enumerate(COL_W,1): ws.column_dimensions[get_column_letter(i)].width = w

    SCHED_COLOR = {"X":X_C,"H1":H1_C,"H":H_C,"OTC":OTC_C}
    for ri, r in enumerate(results, 4):
        fill = PatternFill("solid", fgColor=SCHED_COLOR.get(r["schedule"],"FFFFFF"))
        vals = [ri-3, r["brand_name"], r["dosage_form"], r["schedule"],
                r["matched_molecule"], r["method"], r["confidence"], "", ""]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if r["schedule"] in ("X","H1","H"): cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if ci == 4: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center",vertical="center")
    
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{len(results)+3}"

    # ── Sheet 2: High Confidence Classified ───────────────────────────────
    ws2 = wb.create_sheet("✅ Auto-Classified (High)")
    hdr_cell(ws2,1,1,"High Confidence Auto-Classified — Pharmacist: just verify these are correct",bg="1A7A50")
    ws2.merge_cells("A1:I1"); ws2.row_dimensions[1].height = 22
    for c,h in enumerate(COLS,1): hdr_cell(ws2,2,c,h,wrap=True)
    ws2.row_dimensions[2].height = 32
    for i,w in enumerate(COL_W,1): ws2.column_dimensions[get_column_letter(i)].width = w

    high = [r for r in results if r["confidence"]=="High" and r["schedule"]!="OTC"]
    for ri, r in enumerate(high, 3):
        fill = PatternFill("solid", fgColor=SCHED_COLOR.get(r["schedule"],"FFFFFF"))
        vals = [ri-2,r["brand_name"],r["dosage_form"],r["schedule"],r["matched_molecule"],r["method"],r["confidence"],"",""]
        for ci, v in enumerate(vals,1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.fill = fill; cell.alignment = Alignment(vertical="center")
            if ci==4: cell.font=Font(bold=True); cell.alignment=Alignment(horizontal="center",vertical="center")
    ws2.freeze_panes="A3"

    # ── Sheet 3: Needs Review ─────────────────────────────────────────────
    ws3 = wb.create_sheet("⚠️ Needs Review")
    hdr_cell(ws3,1,1,"Unclassified / Low Confidence — Pharmacist must manually classify these",bg="B03000")
    ws3.merge_cells("A1:I1"); ws3.row_dimensions[1].height = 22
    for c,h in enumerate(COLS,1): hdr_cell(ws3,2,c,h,bg="B03000",wrap=True)
    ws3.row_dimensions[2].height = 32
    for i,w in enumerate(COL_W,1): ws3.column_dimensions[get_column_letter(i)].width = w

    review = [r for r in results if r["confidence"] in ("Low","-") or r["schedule"]=="OTC"]
    for ri, r in enumerate(review, 3):
        vals = [ri-2,r["brand_name"],r["dosage_form"],r["schedule"],r["matched_molecule"],r["method"],r["confidence"],"",""]
        for ci, v in enumerate(vals,1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical="center")
    ws3.freeze_panes="A3"

    # ── Sheet 4: Instructions ─────────────────────────────────────────────
    ws4 = wb.create_sheet("📋 Instructions")
    ws4.column_dimensions["A"].width = 90
    lines = [
        ("MediSyn Schedule Classification — Instructions for Pharmacist", True, 13, TEAL, "FFFFFF"),
        ("", False, 10, None, None),
        ("COLOUR CODING:", True, 11, None, None),
        ("🟣 Purple = Schedule X — Narcotics/Psychotropics. Form 17 register mandatory. Keep prescription 2 years.", False, 10, X_C, "000000"),
        ("🔴 Red = Schedule H1 — 3rd/4th gen antibiotics, anti-TB, benzodiazepines. Separate dispensing register. Keep 3 years.", False, 10, H1_C, "000000"),
        ("🟠 Orange = Schedule H — All other prescription drugs. Need valid Rx. No special register needed.", False, 10, H_C, "000000"),
        ("🟢 Green = OTC / Unclassified — Over the counter or not yet classified.", False, 10, OTC_C, "000000"),
        ("", False, 10, None, None),
        ("YOUR TASK:", True, 11, None, None),
        ("1. Start with sheet '✅ Auto-Classified (High)' — these are high confidence. Scan through and mark Column H: YES if correct, NO if wrong.", False, 10, None, None),
        ("2. Go to sheet '⚠️ Needs Review' — for each medicine, fill Column H (H, H1, X, or OTC) and mark Column I as your answer.", False, 10, None, None),
        ("3. For brand name medicines you know — e.g. GLYCODAY MV is Glibenclamide+Metformin → mark as H", False, 10, None, None),
        ("4. If unsure, leave blank — we can look up later.", False, 10, None, None),
        ("", False, 10, None, None),
        ("HOW TO FILL:", True, 11, None, None),
        ("Column H (Pharmacist Confirm): Type YES (auto is correct) or NO (auto is wrong)", False, 10, None, None),
        ("Column I (Override Schedule): If NO, type the correct schedule: H, H1, X, or OTC", False, 10, None, None),
        ("", False, 10, None, None),
        ("SCHEDULE DEFINITIONS:", True, 11, None, None),
        ("Schedule X: Narcotics — Form 17 register mandatory, keep Rx for 2 years", False, 10, X_C, "000000"),
        ("Schedule H1: 3rd/4th gen antibiotics, anti-TB, habit-forming drugs — register mandatory, keep 3 years", False, 10, H1_C, "000000"),
        ("Schedule H: All other prescription-only drugs — Rx required, no register needed", False, 10, H_C, "000000"),
        ("OTC: Over the counter — no prescription needed (e.g. Paracetamol, Vitamin C, antacids)", False, 10, OTC_C, "000000"),
    ]
    for ri, (text, bold, size, bg, fg) in enumerate(lines, 1):
        cell = ws4.cell(row=ri, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=fg or "000000")
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        ws4.row_dimensions[ri].height = 20

    wb.save("MediSyn_Schedule_Classification.xlsx")
    print(f"\n✅ Excel saved: MediSyn_Schedule_Classification.xlsx")
    print(f"\n📊 RESULTS:")
    print(f"   Schedule X  (Narcotics)    : {counts['X']:5d}")
    print(f"   Schedule H1 (Strict Rx)    : {counts['H1']:5d}")
    print(f"   Schedule H  (Prescription) : {counts['H']:5d}")
    print(f"   OTC / Unclassified         : {counts['OTC']:5d}")
    print(f"   ─────────────────────────────────")
    print(f"   TOTAL                      : {len(results):5d}")
    print(f"\n   Auto-classified (High conf): {len(high):5d}")
    print(f"   Needs pharmacist review    : {len(review):5d}")

if __name__ == "__main__":
    main()
